"""
Port Python de official_csv_adapter.js
Convierte filas de CSV/Excel al payload estándar del validador.
"""
import re
import unicodedata
from pathlib import Path


# ─── Mapeo de departamento por primer dígito del código territorial ──
_DEPT_MAP = {
    "1": "La Paz", "2": "Cochabamba", "3": "Santa Cruz",
    "4": "Oruro",  "5": "Potosi",     "6": "Tarija",
    "7": "Chuquisaca", "8": "Beni",   "9": "Pando",
}


def get_department_name(codigo_territorial: int) -> str:
    return _DEPT_MAP.get(str(codigo_territorial)[:1], "Desconocido")


# ─── Normalización de claves ──────────────────────────────────────
def normalize_key(key: str) -> str:
    key = str(key or "").strip()
    key = unicodedata.normalize("NFD", key)
    key = "".join(c for c in key if unicodedata.category(c) != "Mn")
    key = key.lower().replace(" ", "_")
    key = re.sub(r"[^\w]", "_", key)
    return key


def normalize_row(row: dict) -> dict:
    return {
        normalize_key(k): (v.strip() if isinstance(v, str) else v)
        for k, v in row.items()
    }


def pick(row: dict, aliases: list, default=""):
    for alias in aliases:
        val = row.get(normalize_key(alias))
        if val is not None and val != "":
            return val
    return default


def to_number(value) -> float:
    if value is None or value == "":
        return 0.0
    text = str(value).strip().replace(",", ".").replace(" ", "")
    try:
        return float(text)
    except (ValueError, OverflowError):
        return 0.0


def to_integer(value) -> int:
    return int(to_number(value))


def normalize_long_code(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return ""
    if re.match(r"^\d+$", text):
        return text
    if re.match(r"^\d+\.0+$", text):
        return text.split(".")[0]
    # Scientific notation (e.g. 1.01011e+13)
    if re.match(r"^[+-]?\d+(\.\d+)?[eE][+-]?\d+$", text):
        try:
            n = float(text)
            if n == n and abs(n) < 1e18:  # not NaN / not overflow
                return str(int(n))
        except (ValueError, OverflowError):
            pass
    return re.sub(r"[^\d]", "", text)


def detect_observation_type(text: str) -> str:
    v = str(text or "").lower().strip()
    v = unicodedata.normalize("NFD", v)
    v = "".join(c for c in v if unicodedata.category(c) != "Mn")
    if not v:
        return "SIN_OBSERVACION"
    if "inconsistencia aritmetica" in v:
        return "INCONSISTENCIA_ARITMETICA"
    if "papeletas no autorizadas" in v:
        return "PAPELETAS_NO_AUTORIZADAS"
    if "fecha incorrecta" in v:
        return "FECHA_INCORRECTA"
    if "delegado" in v:
        return "AUSENCIA_DELEGADOS"
    if "apertura" in v or "cierre" in v:
        return "FALTA_DATOS_APERTURA_CIERRE"
    if "tachadura" in v or "enmienda" in v or "corre y vale" in v:
        return "ERROR_TRANSCRIPCION"
    if "firma" in v or "huella" in v:
        return "FALTA_FIRMA_HUELLA"
    if "anulado" in v:
        return "ACTA_ANULADA"
    if ("formulario" in v and "no aprobado" in v) or "documentos no aprobados" in v:
        return "FORMULARIO_NO_OFICIAL"
    if "borrado" in v:
        return "DATOS_BORRADOS"
    if "duplicado" in v:
        return "ACTA_DUPLICADA"
    if "clon" in v or "conado" in v:
        return "ACTA_CLONADA"
    if "mesa" in v and "no existe" in v:
        return "MESA_NO_EXISTE"
    if "mesa" in v and "lugar distinto" in v:
        return "MESA_EN_LUGAR_DISTINTO"
    if "cambiado" in v or "cambio" in v:
        return "DATOS_CAMBIADOS"
    return "OBSERVACION_GENERAL"


_CRITICAL_TYPES = frozenset({
    "ACTA_ANULADA", "FORMULARIO_NO_OFICIAL", "PAPELETAS_NO_AUTORIZADAS",
    "DATOS_BORRADOS", "ACTA_DUPLICADA", "ACTA_CLONADA", "MESA_NO_EXISTE",
    "MESA_EN_LUGAR_DISTINTO", "DATOS_CAMBIADOS", "FALTA_FIRMA_HUELLA",
    "FECHA_INCORRECTA", "ERROR_TRANSCRIPCION", "AUSENCIA_DELEGADOS",
    "FALTA_DATOS_APERTURA_CIERRE", "INCONSISTENCIA_ARITMETICA",
})


def build_official_payload(raw_row: dict, row_number: int = 1, default_user_id: int = 1) -> dict:
    row = normalize_row(raw_row)

    codigo_acta = normalize_long_code(pick(row, [
        "CodigoActa", "codigoacta", "codigo_acta", "nro_acta", "NroActa", "acta",
    ]))
    nro_mesa = to_integer(pick(row, [
        "NroMesa", "nromesa", "nro_mesa", "numero_mesa", "mesa",
    ]))
    votantes_habilitados = to_integer(pick(row, [
        "VotantesHabilitados", "votanteshabilitados", "votantes_habilitados",
        "nro_votantes", "NroVotantes", "habilitados",
    ]))
    papeletas_anfora = to_integer(pick(row, [
        "PapeletasAnfora", "papeletasanfora", "papeletas_anfora",
    ]))
    papeletas_no_utilizadas = to_integer(pick(row, [
        "PapeltasNoUtilizadas", "PapeletasNoUtilizadas",
        "papeltasnoutilizadas", "papeletas_no_utilizadas",
    ]))

    partido1 = to_integer(pick(row, ["P1", "p1", "partido1"]))
    partido2 = to_integer(pick(row, ["P2", "p2", "partido2"]))
    partido3 = to_integer(pick(row, ["P3", "p3", "partido3"]))
    partido4 = to_integer(pick(row, ["P4", "p4", "partido4"]))

    votos_validos_excel = to_integer(pick(row, [
        "VotosValidos", "votosvalidos", "votos_validos",
    ]))
    votos_blancos = to_integer(pick(row, [
        "VotosBlancos", "votosblancos", "votos_blancos",
    ]))
    votos_nulos = to_integer(pick(row, [
        "VotosNulos", "votosnulos", "votos_nulos",
    ]))

    obs_principal = str(pick(row, [
        "Observaciones", "observaciones", "observacion",
        "transcripciones", "transcripcion",
    ], "")).strip()
    obs_extra = str(pick(row, ["__EMPTY", "Unnamed: 13", "unnamed_13", "empty"], "")).strip()
    observaciones = " | ".join(filter(None, [obs_principal, obs_extra]))

    apertura_hora     = to_integer(pick(row, ["AperturaHora",    "aperturahora",    "apertura_hora"]))
    apertura_minutos  = to_integer(pick(row, ["AperturaMinutos", "aperturaminutos", "apertura_minutos"]))
    cierre_hora       = to_integer(pick(row, ["CierreHora",      "cierrehora",      "cierre_hora"]))
    cierre_minutos    = to_integer(pick(row, ["CierreMinutos",   "cierreminutos",   "cierre_minutos"]))

    votos_validos_calculados = partido1 + partido2 + partido3 + partido4
    total_votos = votos_validos_excel + votos_blancos + votos_nulos

    codigo_territorial = int(codigo_acta[:5]) if len(codigo_acta) >= 5 and codigo_acta.isdigit() else 0
    codigo_recinto     = codigo_acta[:10] if len(codigo_acta) >= 10 else codigo_acta[:8] if len(codigo_acta) >= 8 else ""
    nro_mesa_desde_acta = int(codigo_acta[-3:]) if len(codigo_acta) >= 3 and codigo_acta.isdigit() else 0

    tipo_observacion    = detect_observation_type(observaciones)
    tiene_observacion   = tipo_observacion != "SIN_OBSERVACION"
    requiere_revision   = tiene_observacion or tipo_observacion in _CRITICAL_TYPES
    estado_acta         = "OBSERVADA_PENDIENTE_REVISION" if requiere_revision else "VALIDA"

    try:
        codigo_mesa_int = int(codigo_acta) if codigo_acta.isdigit() else 0
    except (ValueError, OverflowError):
        codigo_mesa_int = 0

    # Recinto_id = primeros 8 dígitos del codigo_acta
    try:
        recinto_id = int(codigo_acta[:8]) if len(codigo_acta) >= 8 and codigo_acta.isdigit() else 0
    except (ValueError, OverflowError):
        recinto_id = 0

    return {
        "row_number":           row_number,
        "nro_acta":             codigo_acta,
        "codigo_territorial":   codigo_territorial,
        "codigo_recinto":       codigo_recinto,
        "recinto_id":           recinto_id,
        "codigo_mesa":          codigo_mesa_int,
        "nro_mesa":             nro_mesa,
        "nro_mesa_desde_acta":  nro_mesa_desde_acta,
        "nro_votantes":         votantes_habilitados,
        "papeletas_anfora":     papeletas_anfora,
        "papeletas_no_utilizadas": papeletas_no_utilizadas,
        "votos": {
            "partido1":                partido1,
            "partido2":                partido2,
            "partido3":                partido3,
            "partido4":                partido4,
            "votos_blancos":           votos_blancos,
            "votos_nulos":             votos_nulos,
            "votos_validos":           votos_validos_excel,
            "votos_validos_calculados": votos_validos_calculados,
            "total_votos":             total_votos,
        },
        "registrado_por":         default_user_id,
        "transcripcion":          observaciones,
        "tipo_observacion":       tipo_observacion,
        "requiere_revision_humana": requiere_revision,
        "estado_acta":            estado_acta,
        "apertura":               {"hora": apertura_hora,    "minutos": apertura_minutos},
        "cierre":                 {"hora": cierre_hora,      "minutos": cierre_minutos},
        "origen":                 "EXCEL_OFICIAL_AUTOMATIZADO",
    }


def read_data_file(file_path: str) -> list[dict]:
    """Lee CSV o Excel y retorna lista de dicts."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {file_path}")

    suffix = path.suffix.lower()

    if suffix == ".csv":
        import csv
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return [row for row in reader]

    if suffix in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws_name = "Transcripciones" if "Transcripciones" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[ws_name]
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h or "").strip() for h in next(rows_iter)]
        result = []
        for row in rows_iter:
            if any(v is not None and str(v).strip() != "" for v in row):
                result.append(dict(zip(headers, row)))
        wb.close()
        return result

    raise ValueError(f"Formato no soportado: {suffix}")
